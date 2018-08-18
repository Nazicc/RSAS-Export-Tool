import os
import re
import sys
import html
import queue
import shutil
import zipfile
import datetime
import openpyxl
import images_ico
import PyQt5.sip
from PyQt5.QtGui import QIcon
from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class Vul_re(object):
    def __init__(self):
        super(Vul_re, self).__init__()
        self.vul_list_re = '<python>ip<python>.*?<python>host<python>.*?<td valign="top".*?<th width="120">IP地址</th>.*?<td>(.*?)</td>.*?</td>.*?<python>host</python>.*?<python>vul_list<python>(.*?)<python>vul_list</python>.*?<python>ip</python>'
        self.vul_ip_re = '(<python>ip<python>.*?<python>ip</python>)'
        self.vul_detail_re = '<python>vul_detail<python>(.*?)<python>vul_detail</python>'
        self.vul_details_re = '<python>vul_details<python>(.*?)<python>vul_details</python>'

        self.danger_re = '<span class="level_danger_(.*?)".*?table_\d_(\d+).*?>(.*?)</span>'
        self.title_re = '<python>title<python>(.*?)<python>title</python>'
        self.time_re = '<python>host<python>.*?(\d+-\d+-\d+).*?<python>host</python>'
        self.other_re = '<td class="vul_port">(.*?)</td>.*?<td>(.*?)</td>.*?<td>(.*?)</td>.*?<td>.*?<ul>(.*?)</ul>'

class File_re(object):
    def __init__(self):
        super(File_re, self).__init__()
        self.file_re = '.*?.zip'
        self.uzip_re = '.*?.html'
        self.all_title_re = '<th width="120">任务名称</th>.*?<td>(.*?)</td>'
        self.vul_list_re = '(<table id="vuln_list" class="report_table">.*?</table>)'
        self.vul_detail_re = '(<div id="vul_detail">.*?</div>)'
        self.vul_details_re = '(<tr class="solution.*?">.*?<td>.*?<table class="report_table plumb".*?>.*?</table>.*?</td>.*?</tr>)'
        self.host_re = '(<td valign="top" style="width:50%;">.*?<table class="report_table plumb">.*?<tbody>.*?<th width="120">IP地址</th>.*?</tbody>.*?</table></td>)'

class Vul_content(object):
    def __init__(self,vul_re):
        super(Vul_content, self).__init__()
        self.vul_ip_content = re.findall(vul_re.vul_ip_re,htmlcont,re.S|re.M)
        self.vul_detail_content = re.findall(vul_re.vul_detail_re,htmlcont,re.S|re.M)

class Solve_re(object):
    def __init__(self):
        super(Solve_re, self).__init__()
        self.solve_re = '<th width="100">解决办法</th>.*?<td>(.*?)</td>'
        self.describe_re = '<tr class="solution.*?table_\d_(\d+).*?<th width="100">详细描述</th>.*?<td>(.*?)</td>'
        self.cve_re = '<th width="100">CVE编号</th>.*?<td><a target=.*?>(.*?)</a>.*?</td>'

class Other(object):
    def __init__(self, vul_re, all_vuln_list):
        super(Other, self).__init__()
        self.all_other = re.findall(vul_re.other_re,all_vuln_list,re.S|re.M)

class Danger(object):
    def __init__(self, vul_re, other):
        super(Danger, self).__init__()
        self.danger_coneent = re.findall(vul_re.danger_re,other,re.S|re.M)

class Solve(object):
    def __init__(self, solve, all_vul_details):
        super(Solve, self).__init__()
        self.solve_plumb = re.findall(solve.solve_re,all_vul_details,re.S|re.M)
        self.describe_plumb = re.findall(solve.describe_re,all_vul_details,re.S|re.M)
        self.cve_plumb = re.findall(solve.cve_re,all_vul_details,re.S|re.M)

class Port_File_re(object):
    def __init__(self):
        super(Port_File_re, self).__init__()
        self.file_re = '.*?.zip'
        self.uzip_re = '.*?.html'
        self.all_title_re = '<th width="120">任务名称</th>.*?<td>(.*?)</td>'
        self.host_re = '<th width="120">IP地址</th>.*?<td>(\d+.\d+.\d+.\d+)</td>.*?<th>扫描起始时间</th>.*?<td>(\d+-\d+-\d+).*?</td>.*?<thead>.*?<th>端口</th>.*?<th>协议</th>.*?<th>服务</th>.*?<th>状态</th>.*?</thead>.*?<tbody>(.*?)</tbody>'
        self.port_re = '<tr class=".*?<td>(.*?)</td>.*?<td>(.*?)</td>.*?<td>(.*?)</td>.*?<td>(.*?)</td>.*?</tr>'
        self.http_re = '.*?http.*?'
        self.https_re = '.*?https.*?'
        self.www_re = '.*?www.*?'

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):

        try:
            with open('set.ini') as set_ini:
                name_ini = set_ini.readlines()[:1][0].split(':')[1].strip(' \n')
            with open('set.ini') as set_ini:
                company_ini = set_ini.readlines()[1:2][0].split(':')[1].strip(' \n')
        except Exception as e:
            QtWidgets.QMessageBox.information(MainWindow, "提示", "找不到配置文件，请查看使用说明！", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No ,  QtWidgets.QMessageBox.Yes )

        font = QtGui.QFont()
        font.setFamily("宋体")
        font.setPointSize(10)
        MainWindow.setFont(font)
        #定义程序的标题
        MainWindow.setWindowTitle('RSAS漏洞数据导出工具1.5')
        #设定程序的最大分辨率，禁止最大化、拖动窗口
        MainWindow.setFixedSize(520, 310)
        #设置图标
        MainWindow.setWindowIcon(QIcon(':/favicon.ico'))
        #获取显示器的分辨率
        screen = QtWidgets.QDesktopWidget().screenGeometry()
        #获取程序的宽和高
        size = MainWindow.geometry()
        #实现在屏幕中间显示程序
        MainWindow.move((screen.width() - size.width())/2, (screen.height() - size.height())/2)

        #这是底部的状态栏
        MainWindow.status = MainWindow.statusBar()
        MainWindow.status.showMessage("检查人员：%s  所属公司：%s" % (name_ini,company_ini))

        #这是一个框架，用来固定按钮用的
        self.formLayoutWidget = QtWidgets.QWidget(MainWindow)
        self.formLayoutWidget.setGeometry(QtCore.QRect(10, 10, 411, 54))
        self.formLayoutWidget.setObjectName("formLayoutWidget")
        self.formLayout = QtWidgets.QFormLayout(self.formLayoutWidget)
        self.formLayout.setContentsMargins(0, 0, 0, 0)
        self.formLayout.setObjectName("formLayout")
        #文字：原始报告路径
        self.input_label = QtWidgets.QLabel(self.formLayoutWidget)
        self.input_label.setObjectName("input_label")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.input_label)
        #文字：输出报告路径
        self.output_label = QtWidgets.QLabel(self.formLayoutWidget)
        self.output_label.setObjectName("output_label")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.LabelRole, self.output_label)
        #原始报告路径后边的文本框
        self.input_lineEdit = QtWidgets.QLineEdit(self.formLayoutWidget)
        self.input_lineEdit.setObjectName("input_lineEdit")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.FieldRole, self.input_lineEdit)
        #输出报告路径后边的文本框
        self.output_lineEdit = QtWidgets.QLineEdit(self.formLayoutWidget)
        self.output_lineEdit.setObjectName("output_lineEdit")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.FieldRole, self.output_lineEdit)
        #框架的结束部分
        spacerItem = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.formLayout.setItem(1, QtWidgets.QFormLayout.LabelRole, spacerItem)
        #这玩意就是打开路径按钮的框架
        self.start_verticalLayoutWidget = QtWidgets.QWidget(MainWindow)
        self.start_verticalLayoutWidget.setGeometry(QtCore.QRect(423, 2, 91, 71))
        self.start_verticalLayoutWidget.setObjectName("start_verticalLayoutWidget")
        self.start_verticalLayout = QtWidgets.QVBoxLayout(self.start_verticalLayoutWidget)
        self.start_verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.start_verticalLayout.setObjectName("start_verticalLayout")
        #这是原始报告路径后边的文本框后边的打开路径按钮
        self.input_Button = QtWidgets.QPushButton(self.start_verticalLayoutWidget)
        self.input_Button.setObjectName("input_Button")
        self.start_verticalLayout.addWidget(self.input_Button)
        self.input_Button.clicked.connect(self.input_Button_click)
        #这是输出报告路径后边的文本框后边的打开路径按钮
        self.output_Button = QtWidgets.QPushButton(self.start_verticalLayoutWidget)
        self.output_Button.setObjectName("output_Button")
        self.start_verticalLayout.addWidget(self.output_Button)
        self.output_Button.clicked.connect(self.output_Button_click)
        #这又是一个框架，固定用的
        self.horizontalLayoutWidget = QtWidgets.QWidget(MainWindow)
        self.horizontalLayoutWidget.setGeometry(QtCore.QRect(10, 75, 411, 21))
        self.horizontalLayoutWidget.setObjectName("horizontalLayoutWidget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.horizontalLayoutWidget)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        #文字：选择导出数据
        self.data_label = QtWidgets.QLabel(self.horizontalLayoutWidget)
        self.data_label.setObjectName("data_label")
        self.horizontalLayout.addWidget(self.data_label)
        #复选框：高危
        self.hight_checkBox = QtWidgets.QCheckBox(self.horizontalLayoutWidget)
        self.hight_checkBox.setObjectName("hight_checkBox")
        self.horizontalLayout.addWidget(self.hight_checkBox)
        #复选框：中危
        self.middle_checkBox = QtWidgets.QCheckBox(self.horizontalLayoutWidget)
        self.middle_checkBox.setObjectName("middle_checkBox")
        self.horizontalLayout.addWidget(self.middle_checkBox)
        #复选框：低危
        self.low_checkBox = QtWidgets.QCheckBox(self.horizontalLayoutWidget)
        self.low_checkBox.setObjectName("low_checkBox")
        self.horizontalLayout.addWidget(self.low_checkBox)
        #复选框：端口
        self.port_checkBox = QtWidgets.QCheckBox(self.horizontalLayoutWidget)
        self.port_checkBox.setObjectName("port_checkBox")
        self.horizontalLayout.addWidget(self.port_checkBox)
        #复选框：网站
        self.web_checkBox = QtWidgets.QCheckBox(self.horizontalLayoutWidget)
        self.web_checkBox.setObjectName("web_checkBox")
        self.horizontalLayout.addWidget(self.web_checkBox)
        #这又是一个框架
        self.end_verticalLayoutWidget = QtWidgets.QWidget(MainWindow)
        self.end_verticalLayoutWidget.setGeometry(QtCore.QRect(423, 69, 91, 31))
        self.end_verticalLayoutWidget.setObjectName("end_verticalLayoutWidget")
        self.end_verticalLayout = QtWidgets.QVBoxLayout(self.end_verticalLayoutWidget)
        self.end_verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.end_verticalLayout.setObjectName("end_verticalLayout")
        #按钮：开始导出
        self.start_Button = QtWidgets.QPushButton(self.end_verticalLayoutWidget)
        self.start_Button.setObjectName("start_Button")
        self.start_Button.clicked.connect(self.start_Button_click)
        #框架结尾
        self.end_verticalLayout.addWidget(self.start_Button)
        #文字：详细输出日志
        self.log_label = QtWidgets.QLabel(MainWindow)
        self.log_label.setGeometry(QtCore.QRect(10, 101, 91, 21))
        self.log_label.setObjectName("log_label")
        #详细输出日志的文本框
        self.log_textEdit = QtWidgets.QTextEdit(MainWindow)
        self.log_textEdit.setGeometry(QtCore.QRect(10, 120, 501, 171))
        self.log_textEdit.setObjectName("log_textEdit")

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        self.input_label.setText(_translate("MainWindow", "原始报告路径："))
        self.output_label.setText(_translate("MainWindow", "输出报告路径："))
        self.input_Button.setText(_translate("MainWindow", "打开路径"))
        self.output_Button.setText(_translate("MainWindow", "打开路径"))
        self.data_label.setText(_translate("MainWindow", "选择导出数据："))
        self.hight_checkBox.setText(_translate("MainWindow", "高危"))
        self.middle_checkBox.setText(_translate("MainWindow", "中危"))
        self.low_checkBox.setText(_translate("MainWindow", "低危"))
        self.port_checkBox.setText(_translate("MainWindow", "端口"))
        self.web_checkBox.setText(_translate("MainWindow", "网站"))
        self.start_Button.setText(_translate("MainWindow", "开始导出"))
        self.log_label.setText(_translate("MainWindow", "详细输出日志："))

    #原始报告路径的按钮
    def input_Button_click(self):
        self.input_Button_cent = QtWidgets.QFileDialog.getExistingDirectory(MainWindow)
        self.input_lineEdit.setText(self.input_Button_cent)
    
    #输出报告路径的按钮
    def output_Button_click(self):
        self.output_Button_cent = QtWidgets.QFileDialog.getExistingDirectory(MainWindow)
        self.output_lineEdit.setText(self.output_Button_cent)

    #开始导出的按钮
    def start_Button_click(self):
        #复选框事件
        self.hight_status = self.hight_checkBox.isChecked()
        self.middle_status = self.middle_checkBox.isChecked()
        self.low_status = self.low_checkBox.isChecked()
        self.port_status = self.port_checkBox.isChecked()
        self.web_status = self.web_checkBox.isChecked()

        try:
            folder_end = self.output_Button_cent
            folder_start = self.input_Button_cent
        except Exception as e:
            QtWidgets.QMessageBox.information(MainWindow, "提示", "要先设置文件夹！", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No ,  QtWidgets.QMessageBox.Yes )
            return e
        else:
            try:
                with open('set.ini') as set_ini:
                    name_ini = set_ini.readlines()[:1][0].split(':')[1].strip(' \n')
                with open('set.ini') as set_ini:
                    company_ini = set_ini.readlines()[1:2][0].split(':')[1].strip(' \n')
            except Exception as e:
                QtWidgets.QMessageBox.information(MainWindow, "提示", "找不到配置文件，请查看使用说明！", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No ,  QtWidgets.QMessageBox.Yes )
                return e

        try:
            shutil.rmtree('temp')
        except Exception as e:
            pass

        try:
            shutil.rmtree(folder_end+'/汇总-漏洞跟踪表')
        except Exception as e:
            pass

        try:
            shutil.rmtree(folder_end+'/汇总-端口对应关系表')
        except Exception as e:
            pass

        try:
            shutil.rmtree(folder_end+'/汇总-WEB网站')
        except Exception as e:
            pass

        if self.hight_status or self.middle_status or self.low_status == True:
            starttime = datetime.datetime.now()
            self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
            self.log_textEdit.insertPlainText('正在提取数据，这个过程可能很慢，请耐心等待！' + '\n')
            QtWidgets.QApplication.processEvents()

            os.mkdir('temp')
            with open('temp/database.mdb', 'w',encoding='gb18030') as content:
                content.write('')

            dirList = os.listdir(folder_start)
            for name in dirList:
                all_file_name = re.findall(File_re().file_re,name)
                for file_name in all_file_name:
                    uzip = zipfile.ZipFile(folder_start+'/'+file_name)
                    for uzip_content in uzip.namelist():
                        all_uzip_content = re.findall(File_re().uzip_re,uzip_content)
                        for all_uzip in all_uzip_content:
                            htmlcont_zip = uzip.open(all_uzip).read().decode('utf8')
                            title = re.findall(File_re().all_title_re,htmlcont_zip,re.S|re.M)
                            for title_content in title:
                                with open('temp/%s.mdb'%title_content, 'a',encoding='gb18030') as content:
                                    content.write('<python>title<python>')
                                    content.write(html.unescape(title_content))
                                    content.write('<python>title</python>\n')

                                with open('temp/database.mdb','a',encoding='gb18030') as content:
                                    content.write('temp/'+html.unescape(title_content)+'.mdb\n')

                            host = re.findall(File_re().host_re,htmlcont_zip,re.S|re.M)
                            for host_content in host:
                                with open('temp/%s.mdb'%title_content, 'a',encoding='gb18030') as content:
                                    content.write('<python>ip<python>\n')
                                    content.write('<python>host<python>\n')
                                    content.write(html.unescape(host_content))
                                    content.write('\n<python>host</python>\n')

                            vul_list = re.findall(File_re().vul_list_re,htmlcont_zip,re.S|re.M)
                            for list_content in vul_list:
                                with open('temp/%s.mdb'%title_content, 'a',encoding='gb18030') as content:
                                    content.write('<python>vul_list<python>\n')
                                    content.write(html.unescape(list_content))
                                    content.write('\n<python>vul_list</python>\n')

                            vul_detail = re.findall(File_re().vul_detail_re,htmlcont_zip,re.S|re.M)
                            for detail_content in vul_detail:
                                with open('temp/%s.mdb'%title_content, 'a',encoding='gb18030') as content:
                                    content.write('<python>vul_detail<python>\n')

                                vul_details = re.findall(File_re().vul_details_re,detail_content,re.S|re.M)
                                for list_details in vul_details:
                                    with open('temp/%s.mdb'%title_content, 'a',encoding='gb18030') as content:
                                        content.write('<python>vul_details<python>\n')
                                        content.write(html.unescape(list_details))
                                        content.write('\n<python>vul_details</python>\n')

                            with open('temp/%s.mdb'%title_content, 'a',encoding='gb18030') as content:
                                content.write('\n<python>vul_detail</python>\n')
                                content.write('<python>ip</python>\n')

                    self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
                    self.log_textEdit.insertPlainText('提取 %s'%file_name + ' 完成！\n')
                    QtWidgets.QApplication.processEvents()

            vul_list = queue.Queue()
            self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
            self.log_textEdit.insertPlainText('数据提取完成，正在生成漏洞跟踪表...\n')
            QtWidgets.QApplication.processEvents()
            os.mkdir(folder_end+'/汇总-漏洞跟踪表')

            vul_re = Vul_re()
            
            with open('temp/database.mdb',encoding='gb18030') as content:
                for zip_content in content:
                    vul_all_list = []
                    vul_all_detail = []

                    zip_cont = zip_content.strip('\n\r')
                    content = open(zip_cont,'r',encoding='gb18030')
                    global htmlcont
                    htmlcont = content.read()
                    content.close()
                    
                    sheet_name =  re.findall(vul_re.title_re,htmlcont,re.S|re.M)[0]
                    sheet_time =  re.findall(vul_re.time_re,htmlcont,re.S|re.M)[0]
                    self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
                    self.log_textEdit.insertPlainText('正在导出 %s\n' % sheet_name)
                    QtWidgets.QApplication.processEvents()

                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = '漏洞数据'
                    ws.column_dimensions['A'].width = 6
                    ws.column_dimensions['B'].width = 13
                    ws.column_dimensions['C'].width = 24
                    ws.column_dimensions['D'].width = 11
                    ws.column_dimensions['E'].width = 11
                    ws.column_dimensions['F'].width = 53
                    ws.column_dimensions['G'].width = 6
                    ws.column_dimensions['H'].width = 6
                    ws.column_dimensions['I'].width = 53
                    ws.column_dimensions['J'].width = 53
                    ws.column_dimensions['K'].width = 15
                    ws.column_dimensions['L'].width = 11
                    ws.column_dimensions['M'].width = 13
                    ws.column_dimensions['N'].width = 13
                    ws.column_dimensions['O'].width = 13
                    ws.column_dimensions['P'].width = 16
                    ws.column_dimensions['Q'].width = 13
                    ws.column_dimensions['S'].width = 9
                    ws.column_dimensions['T'].width = 9
                    ws.column_dimensions['U'].width = 9
                    ws.column_dimensions['V'].width = 9
                    ws.column_dimensions['W'].width = 9
                    ws.column_dimensions['X'].width = 9
                    ws.column_dimensions['Y'].width = 11

                    ws['A3'] = '序号'
                    ws['B3'] = '主机名'
                    ws['C3'] = 'IP地址'
                    ws['D3'] = '漏洞对应端口'
                    ws['E3'] = '漏洞对应服务'
                    ws['F3'] = '漏洞名称'
                    ws['G3'] = '风险分类'
                    ws['H3'] = '风险等级'
                    ws['I3'] = '整改建议'
                    ws['J3'] = '漏洞描述'
                    ws['K3'] = '漏洞CVE编号'
                    ws['L3'] = '漏洞对应协议'
                    ws['M3'] = '是否已备案'
                    ws['N3'] = '是否已整改'
                    ws['O3'] = '%s核实是否已经整改' % company_ini
                    ws['P3'] = '漏洞复核时间'
                    ws['Q3'] = '备注'
                    ws['S1'] = '高风险总数'
                    ws['T1'] = '中风险总数'
                    ws['U1'] = '修补高风险数'
                    ws['V1'] = '修补中风险数'
                    ws['W1'] = '高风险修补率'
                    ws['X1'] = '中风险修补率'
                    ws['Y1'] = '总修补率'
                    ws['S2'].value = '=COUNTIF(H:H,"高")'
                    ws['T2'].value = '=COUNTIF(H:H,"中")'
                    ws['U2'].value = '=COUNTIFS(H:H,"高",O:O,"已完成整改")'
                    ws['V2'].value = '=COUNTIFS(H:H,"中",O:O,"已完成整改")'
                    ws['W2'].value = '=IF(S2=0,"无高风险漏洞",TEXT(U2/S2,"0.00%"))'
                    ws['X2'].value = '=IF(T2=0,"无中风险漏洞",TEXT(V2/T2,"0.00%"))'
                    ws['Y2'].value = '=TEXT((U2+V2)/(S2+T2),"0.00%")'

                    ws['A1'] = '高中风险漏洞整改情况跟踪表'
                    ws.merge_cells('A1:Q1')
                    ws['A2'] = '系统名称'
                    ws.merge_cells('A2:B2')
                    ws['C2'] = ''
                    ws.merge_cells('C2:F2')
                    ws['G2'] = '管理员'
                    ws.merge_cells('G2:H2')
                    ws['K2'] = '检查人员'
                    ws['L2'] = '%s'% name_ini 
                    ws.merge_cells('L2:M2')
                    ws['O2'] = '检查时间'
                    ws['P2'] = '%s' % sheet_time
                    ws.merge_cells('P2:Q2')

                    # 样式
                    font = Font(size=10, name='宋体')
                    thin = Side(border_style="thin")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    # 对齐
                    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    #设置第三行的格式
                    title_font = Font(size=12, bold=True, name='宋体', color= "ff0000")
                    ws.row_dimensions[3].height = 45
                    for title_style in ws['A3:L3']:
                        for title_cell in title_style:
                            title_cell.font = title_font
                            title_cell.border = border
                            title_cell.alignment = alignment

                    #设置第三行的格式
                    title_end_font = Font(size=12, bold=True, name='宋体', color= "0000FF")
                    ws.row_dimensions[3].height = 45
                    for title_end_style in ws['M3:Q3']:
                        for title_cell in title_end_style:
                            title_cell.font = title_end_font
                            title_cell.border = border
                            title_cell.alignment = alignment

                    #设置第一行的格式
                    one_font = Font(size=20, bold=True, name='宋体')
                    ws.row_dimensions[1].height = 45
                    for one_style in ws['A1:L1']:
                        for one_cell in one_style:
                            one_cell.font = one_font
                            one_cell.border = border
                            one_cell.alignment = alignment
                            one_cell.fill = openpyxl.styles.PatternFill(fill_type='solid',fgColor="008000")

                    #设置第二行的格式
                    two_font = Font(size=14, bold=True, name='宋体')
                    ws.row_dimensions[2].height = 20
                    for one_style in ws['A2:Q2']:
                        for one_cell in one_style:
                            one_cell.font = two_font
                            one_cell.border = border
                            one_cell.alignment = alignment
                            one_cell.fill = openpyxl.styles.PatternFill(fill_type='solid',fgColor="008000")

                    #设置统计的格式
                    count_font = Font(size=12, bold=True, name='宋体')
                    for count_style in ws['S1:Y2']:
                        for count_cell in count_style:
                            count_cell.font = count_font
                            count_cell.border = border
                            count_cell.alignment = alignment

                    count_red = openpyxl.styles.PatternFill(fill_type='solid',fgColor="FF5353")
                    count_yellow = openpyxl.styles.PatternFill(fill_type='solid',fgColor="FCCC2C")
                    ws['S1'].fill = count_red
                    ws['U1'].fill = count_red
                    ws['W1'].fill = count_red
                    ws['T1'].fill = count_yellow
                    ws['V1'].fill = count_yellow
                    ws['X1'].fill = count_yellow
                    ws['Y1'].fill = openpyxl.styles.PatternFill(fill_type='solid',fgColor="008000")

                    vul_content = Vul_content(vul_re)

                    for all_vul_ip in vul_content.vul_ip_content:
                        vul_list_content = re.findall(vul_re.vul_list_re,all_vul_ip,re.S|re.M)
                        for all_vul_list in vul_list_content:
                            for other in Other(vul_re,all_vul_list[1]).all_other:
                                for danger in Danger(vul_re,other[3]).danger_coneent:
                                    vul_all_list.append([danger[1],all_vul_list[0],danger[2],danger[0].replace('low','低').replace('middle','中').replace('high','高'),other[0],other[1],other[2]])

                    for all_vul_detail in vul_content.vul_detail_content:
                        vul_details_content = re.findall(vul_re.vul_details_re,all_vul_detail,re.S|re.M)
                        for all_vul_details in vul_details_content:
                            vul_detail = Solve(Solve_re(),all_vul_details)
                            for solve,describe in zip(vul_detail.solve_plumb,vul_detail.describe_plumb):
                                cve = vul_detail.cve_plumb
                                if cve:
                                    pass
                                else:
                                    cve = ['漏洞暂无CVE编号']
                                vul_all_detail.append([describe[0],re.sub('\s{2,}','\n',html.unescape(re.sub('\s{2,}','',solve)).replace('<br/>','\n')),re.sub('\s{2,}','\n',html.unescape(re.sub('\s{2,}','',describe[1])).replace('<br/>','\n')),cve[0]])

                    for line_vul_list in vul_all_list:
                        vul_list.put(line_vul_list)

                    i = 1
                    while not vul_list.empty():
                        wait_list = vul_list.get()
                        for wait_detail in vul_all_detail:
                            if wait_list[0] == wait_detail[0] and self.hight_status == True and wait_list[3] == '高':
                                ws.row_dimensions[i+3].height = 25
                                ws.append([i,'',wait_list[1],wait_list[4],wait_list[6],wait_list[2],'漏洞',wait_list[3],wait_detail[1].strip('\n'),wait_detail[2].strip('\n'),wait_detail[3],wait_list[5]])
                                for row in ws['A%s:Q%s'%(i+3,i+3)]:
                                    for cell in row:
                                        cell.font = font
                                        cell.border = border
                                        cell.alignment = alignment
                                i += 1
                                break

                            if wait_list[0] == wait_detail[0] and self.middle_status == True and wait_list[3] == '中':
                                ws.row_dimensions[i+3].height = 25
                                ws.append([i,'',wait_list[1],wait_list[4],wait_list[6],wait_list[2],'漏洞',wait_list[3],wait_detail[1].strip('\n'),wait_detail[2].strip('\n'),wait_detail[3],wait_list[5]])
                                for row in ws['A%s:Q%s'%(i+3,i+3)]:
                                    for cell in row:
                                        cell.font = font
                                        cell.border = border
                                        cell.alignment = alignment
                                i += 1
                                break

                            if wait_list[0] == wait_detail[0] and self.low_status == True and wait_list[3] == '低':
                                ws.row_dimensions[i+3].height = 25
                                ws.append([i,'',wait_list[1],wait_list[4],wait_list[6],wait_list[2],'漏洞',wait_list[3],wait_detail[1].strip('\n'),wait_detail[2].strip('\n'),wait_detail[3],wait_list[5]])
                                for row in ws['A%s:Q%s'%(i+3,i+3)]:
                                    for cell in row:
                                        cell.font = font
                                        cell.border = border
                                        cell.alignment = alignment
                                i += 1
                                break

                    wb.save(folder_end+'/汇总-漏洞跟踪表/高中风险漏洞跟踪表--%s.xlsx' % sheet_name)
                    del vul_all_list[:]
                    del vul_all_detail[:]
                self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
                self.log_textEdit.insertPlainText('漏洞跟踪表导出完成，保存在 %s/汇总-漏洞跟踪表 目录下。\n'%folder_end)
                QtWidgets.QApplication.processEvents()

            shutil.rmtree('temp')
            endtime = datetime.datetime.now()
            self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
            self.log_textEdit.insertPlainText('导出花时：%s秒...\n\n\n'%(endtime - starttime).seconds)
            QtWidgets.QApplication.processEvents()


        if self.port_status:
            self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
            self.log_textEdit.insertPlainText('正在导出端口，请稍后！\n')
            QtWidgets.QApplication.processEvents()
            starttime = datetime.datetime.now()
            os.mkdir(folder_end+'/汇总-端口对应关系表')
            dirList = os.listdir(folder_start)
            for name in dirList:
                all_file_name = re.findall(Port_File_re().file_re,name)
                for file_name in all_file_name:
                    uzip = zipfile.ZipFile(folder_start+'/'+file_name)
                    i = 1
                    wb = openpyxl.Workbook()
                    ws = wb.active

                    ws.column_dimensions['A'].width = 16.5
                    ws.column_dimensions['B'].width = 16
                    ws.column_dimensions['C'].width = 20
                    ws.column_dimensions['D'].width = 30
                    ws.column_dimensions['E'].width = 25
                    ws.column_dimensions['F'].width = 28
                    ws.column_dimensions['G'].width = 42
                    ws.column_dimensions['H'].width = 17


                    ws.title = '端口数据'
                    ws['A1'] = '设备端口和服务信息表'
                    ws.merge_cells('A1:H1')
                    ws['A2'] = '收集时间'
                    ws.merge_cells('A2:B2')
                    ws.merge_cells('C2:D2')
                    ws['E2'] = '所属系统'
                    ws.merge_cells('F2:H2')
                    ws['A3'] = '填表人'
                    ws['C3'] = name_ini
                    ws.merge_cells('A3:B3')
                    ws.merge_cells('C3:D3')
                    ws['E3'] = '系统责任人'
                    ws.merge_cells('F3:H3')
                    ws['A4'] = 'IP地址'
                    ws['B4'] = '端口'
                    ws['C4'] = '协议'
                    ws['D4'] = '服务'
                    ws['E4'] = '状态'
                    ws['F4'] = '访问权限开放范围'
                    ws['G4'] = '应用说明'
                    ws['H4'] = '备注'

                    # 样式
                    font = Font(size=12, name='宋体')
                    thin = Side(border_style="thin")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    # 对齐
                    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    for excel_style in ws['A2:H4']:
                        for excel_cell in excel_style:
                            excel_cell.font = font
                            excel_cell.border = border
                            excel_cell.alignment = alignment
                    #设置第一行的格式
                    one_font = Font(size=12, bold=True, name='宋体')
                    for one_style in ws['A1:H1']:
                        for one_cell in one_style:
                            one_cell.font = one_font
                            one_cell.border = border
                            one_cell.alignment = alignment

                    for uzip_content in uzip.namelist():
                        all_uzip_content = re.findall(Port_File_re().uzip_re,uzip_content)
                        for all_uzip in all_uzip_content:
                            htmlcont_zip = uzip.open(all_uzip).read().decode('utf8')
                            vul_title = re.findall(Port_File_re().all_title_re,htmlcont_zip,re.S|re.M)
                            for title_content in vul_title:
                                pass

                            vul_host = re.findall(Port_File_re().host_re,htmlcont_zip,re.S|re.M)
                            for host_content in vul_host:
                              for vul_port in re.findall(Port_File_re().port_re,host_content[2],re.S|re.M):
                                ws['C2'] = '%s' % host_content[1]
                                ws.append([host_content[0],vul_port[0].replace(' ','').strip('\n'),vul_port[1].replace(' ','').strip('\n'),vul_port[2].replace(' ','').strip('\n'),vul_port[3].replace(' ','').strip('\n')])
                                for row in ws['A%s:H%s'%(i+4,i+4)]:
                                    for cell in row:
                                        cell.font = font
                                        cell.border = border
                                        cell.alignment = alignment
                                i += 1

                    wb.save(folder_end+'/汇总-端口对应关系表/端口服务对应关系表--%s.xlsx' % title_content)
                    self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
                    self.log_textEdit.insertPlainText('导出 %s'%title_content+' 完成！\n')
                    QtWidgets.QApplication.processEvents()

            endtime = datetime.datetime.now()
            self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
            self.log_textEdit.insertPlainText('所有端口导出完成，保存在 %s/汇总-端口对应关系表 目录下。\n'%folder_end)
            QtWidgets.QApplication.processEvents()
            self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
            self.log_textEdit.insertPlainText('导出花时：%s秒...\n\n\n'%(endtime - starttime).seconds)
            QtWidgets.QApplication.processEvents()

        if self.web_status:
            self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
            self.log_textEdit.insertPlainText('正在WEB网站，请稍后！\n')
            QtWidgets.QApplication.processEvents()
            starttime = datetime.datetime.now()
            os.mkdir(folder_end+'/汇总-WEB网站')
            dirList = os.listdir(folder_start)
            for name in dirList:
                all_file_name = re.findall(Port_File_re().file_re,name)
                for file_name in all_file_name:
                    uzip = zipfile.ZipFile(folder_start+'/'+file_name)
                    x = 1

                    web = openpyxl.Workbook()
                    wes = web.active
                    wes.title = 'WEB网站'

                    wes.column_dimensions['A'].width = 16.5
                    wes.column_dimensions['B'].width = 16
                    wes.column_dimensions['C'].width = 20
                    wes.column_dimensions['D'].width = 30
                    wes.column_dimensions['E'].width = 25
                    wes.column_dimensions['F'].width = 45

                    wes['A1'] = 'IP地址'
                    wes['B1'] = '端口'
                    wes['C1'] = '协议'
                    wes['D1'] = '服务'
                    wes['E1'] = '状态'
                    wes['F1'] = 'WEB网站信息'

                    # 样式
                    font = Font(size=12, name='宋体')
                    thin = Side(border_style="thin")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    # 对齐
                    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    for excel_style in wes['A1:F1']:
                        for excel_cell in excel_style:
                            excel_cell.font = font
                            excel_cell.border = border
                            excel_cell.alignment = alignment

                    for uzip_content in uzip.namelist():
                        all_uzip_content = re.findall(Port_File_re().uzip_re,uzip_content)
                        for all_uzip in all_uzip_content:
                            htmlcont_zip = uzip.open(all_uzip).read().decode('utf8')
                            vul_title = re.findall(Port_File_re().all_title_re,htmlcont_zip,re.S|re.M)
                            for title_content in vul_title:
                                pass

                            vul_host = re.findall(Port_File_re().host_re,htmlcont_zip,re.S|re.M)
                            for host_content in vul_host:
                              for vul_port in re.findall(Port_File_re().port_re,host_content[2],re.S|re.M):
                                vul_web = re.findall(Port_File_re().http_re,vul_port[2].replace(' ','').strip('\n'),re.S|re.M)
                                if vul_web:
                                  wes.append([host_content[0],vul_port[0].replace(' ','').strip('\n'),vul_port[1].replace(' ','').strip('\n'),vul_port[2].replace(' ','').strip('\n'),vul_port[3].replace(' ','').strip('\n'),'http://'+ host_content[0] + ':' + vul_port[0].replace(' ','').strip('\n')])
                                  for row in wes['A%s:E%s'%(x+1,x+1)]:
                                    for cell in row:
                                        cell.font = font
                                        cell.border = border
                                        cell.alignment = alignment
                                    wes['F%s' % (x+1)].font = font
                                    wes['F%s' % (x+1)].border = border
                                  x += 1
                                vul_web = re.findall(Port_File_re().https_re,vul_port[2].replace(' ','').strip('\n'),re.S|re.M)
                                if vul_web:
                                  wes.append([host_content[0],vul_port[0].replace(' ','').strip('\n'),vul_port[1].replace(' ','').strip('\n'),vul_port[2].replace(' ','').strip('\n'),vul_port[3].replace(' ','').strip('\n'),'https://'+ host_content[0] + ':' + vul_port[0].replace(' ','').strip('\n')])
                                  for row in wes['A%s:E%s'%(x+1,x+1)]:
                                    for cell in row:
                                        cell.font = font
                                        cell.border = border
                                        cell.alignment = alignment
                                    wes['F%s' % (x+1)].font = font
                                    wes['F%s' % (x+1)].border = border
                                  x += 1
                                vul_web = re.findall(Port_File_re().www_re,vul_port[2].replace(' ','').strip('\n'),re.S|re.M)
                                if vul_web:
                                  wes.append([host_content[0],vul_port[0].replace(' ','').strip('\n'),vul_port[1].replace(' ','').strip('\n'),vul_port[2].replace(' ','').strip('\n'),vul_port[3].replace(' ','').strip('\n'),'http://'+ host_content[0] + ':' + vul_port[0].replace(' ','').strip('\n')])
                                  for row in wes['A%s:E%s'%(x+1,x+1)]:
                                    for cell in row:
                                        cell.font = font
                                        cell.border = border
                                        cell.alignment = alignment
                                    wes['F%s' % (x+1)].font = font
                                    wes['F%s' % (x+1)].border = border
                                  x += 1

                    web.save(folder_end+'/汇总-WEB网站/WEB网站--%s.xlsx' % title_content)
                    self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
                    self.log_textEdit.insertPlainText('导出 %s'%title_content+' 完成！\n')
                    QtWidgets.QApplication.processEvents()
            endtime = datetime.datetime.now()
            self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
            self.log_textEdit.insertPlainText('所有WEB网站导出导出完成，保存在 %s/汇总-WEB网站 目录下。\n'%folder_end)
            QtWidgets.QApplication.processEvents()
            self.log_textEdit.moveCursor(QtGui.QTextCursor.End)
            self.log_textEdit.insertPlainText('导出花时：%s秒...\n\n\n'%(endtime - starttime).seconds)
            QtWidgets.QApplication.processEvents()

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())